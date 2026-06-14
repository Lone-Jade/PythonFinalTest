import ast
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


@dataclass
class JobShopInstance:
    name: str
    n_jobs: int
    n_machines: int
    n_workers: int
    machines: np.ndarray
    processing_times: np.ndarray

    @property
    def n_operations(self) -> int:
        return self.n_jobs * self.n_machines


INSTANCE_RE = re.compile(r"^\s*(\d+)\s*x\s*(\d+)\s*x\s*(\d+)\s*$", re.IGNORECASE)


def _parse_processing_list(value, n_workers):
    if isinstance(value, str):
        parsed = ast.literal_eval(value.replace("，", ","))
    elif isinstance(value, (list, tuple)):
        parsed = value
    else:
        raise ValueError(f"Bad processing-time cell: {value!r}")
    if len(parsed) != n_workers:
        raise ValueError(f"Expected {n_workers} worker times, got {parsed!r}")
    return [float(x) for x in parsed]


def load_instances(path="basic_data.xlsx", sheet_names=None, skip_bad=True):
    """Parse all benchmark blocks from the workbook.

    Each block has a title like 10x5x3, a Job/Machine header row, then n_jobs
    rows. Machine order is fixed by the data; worker-dependent times are stored
    as lists such as [160,146,104].
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    names = sheet_names or wb.sheetnames
    instances = {}

    for sheet_name in names:
        ws = wb[sheet_name]
        row = 1
        while row <= ws.max_row:
            first = ws.cell(row, 1).value
            if not isinstance(first, str):
                row += 1
                continue
            match = INSTANCE_RE.match(first)
            if not match:
                row += 1
                continue

            n_jobs, n_machines, n_workers = map(int, match.groups())

            header_row = None
            for rr in range(row + 1, min(ws.max_row, row + 8) + 1):
                if str(ws.cell(rr, 1).value).strip().lower() == "job":
                    header_row = rr
                    break
            if header_row is None:
                row += 1
                continue

            first_job_row = header_row + 1
            proc_start_col = None
            for col in range(2 + n_machines, ws.max_column + 1):
                value = ws.cell(first_job_row, col).value
                if isinstance(value, str) and value.strip().startswith("["):
                    proc_start_col = col
                    break
            if proc_start_col is None:
                raise ValueError(f"Cannot find processing-time columns in {sheet_name}:{row}")

            machines = np.zeros((n_jobs, n_machines), dtype=np.int64)
            processing = np.zeros((n_jobs, n_machines, n_workers), dtype=np.float32)

            bad_block = None
            for j in range(n_jobs):
                rr = first_job_row + j
                for op in range(n_machines):
                    machines[j, op] = int(ws.cell(rr, 2 + op).value) - 1
                    try:
                        processing[j, op, :] = _parse_processing_list(
                            ws.cell(rr, proc_start_col + op).value, n_workers
                        )
                    except Exception as exc:
                        cell = ws.cell(rr, proc_start_col + op).coordinate
                        bad_block = ValueError(
                            f"Bad processing time at sheet={sheet_name}, block={first}, "
                            f"job={j + 1}, op={op + 1}, cell={cell}"
                        )
                        bad_block.__cause__ = exc
                        break
                if bad_block is not None:
                    break

            if bad_block is not None:
                if skip_bad:
                    print(f"[WARN] Skip malformed instance: {bad_block}")
                    row = first_job_row + n_jobs
                    continue
                raise bad_block

            key = f"{sheet_name}_{n_jobs}x{n_machines}x{n_workers}"
            if key in instances:
                suffix = 2
                while f"{key}_{suffix}" in instances:
                    suffix += 1
                key = f"{key}_{suffix}"
            instances[key] = JobShopInstance(
                name=key,
                n_jobs=n_jobs,
                n_machines=n_machines,
                n_workers=n_workers,
                machines=machines,
                processing_times=processing,
            )
            row = first_job_row + n_jobs

    return instances


def list_instances(path="basic_data.xlsx"):
    instances = load_instances(path)
    return [
        (name, inst.n_jobs, inst.n_machines, inst.n_workers)
        for name, inst in instances.items()
    ]
