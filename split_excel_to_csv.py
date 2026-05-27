"""将 basic_data.xlsx 按规模拆分为多个 CSV 文件。

Excel 中每个 sheet 是一个规模 (如 6x6, 10x5)，每个 sheet 内可能包含多个子问题
(如 6x6x3, 6x6x2 分别代表不同工人数量)。
"""

import csv
import os
import re

import openpyxl


def parse_sub_problems(ws):
    """解析 worksheet 中的所有子问题，返回 [(name, header_row, data_rows)] 列表。"""
    sub_problems = []
    row_idx = 0
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    while row_idx < len(all_rows):
        row = all_rows[row_idx]
        # 检测子问题标识行，如 "6x6x3", "10x5x3", "20x10x6"
        first_val = row[0] if row else None
        if first_val and isinstance(first_val, str) and re.match(r'^\d+x\d+x\d+$', first_val.strip()):
            name = first_val.strip()
            # 跳过 NoJE = X 行
            row_idx += 1
            # 跳过可能的空行 和 "Processing Time" 标签行
            row_idx = _skip_meta_rows(all_rows, row_idx)

            # 找表头行 (Job, Machine, ...)
            if row_idx < len(all_rows):
                header_row = all_rows[row_idx]
                row_idx += 1
                # 收集数据行直到遇到空行或下一个子问题标识
                data_rows = []
                while row_idx < len(all_rows):
                    r = all_rows[row_idx]
                    first_cell = r[0] if r else None
                    # 遇到下一个子问题标识或空行则停止
                    if first_cell and isinstance(first_cell, str) and re.match(r'^\d+x\d+x\d+$', first_cell.strip()):
                        break
                    if all(v is None for v in r):
                        row_idx += 1
                        break
                    data_rows.append(r)
                    row_idx += 1
                sub_problems.append((name, header_row, data_rows))
                continue
        row_idx += 1

    return sub_problems


def _skip_meta_rows(all_rows, row_idx):
    """跳过 NoJE、空行、Processing Time 等元数据行。"""
    while row_idx < len(all_rows):
        r = all_rows[row_idx]
        first_val = r[0] if r else None
        non_none = [v for v in r if v is not None]
        if not non_none:
            row_idx += 1
            continue
        if first_val and isinstance(first_val, str):
            stripped = first_val.strip()
            if stripped.startswith('NoJE') or stripped.startswith('Processing Time'):
                row_idx += 1
                continue
        # 遇到表头行 (Job 开头) 则停止
        if first_val and isinstance(first_val, str) and first_val.strip() == 'Job':
            break
        # 遇到下一个子问题标识则停止
        if first_val and isinstance(first_val, str) and re.match(r'^\d+x\d+x\d+$', first_val.strip()):
            break
        break
    return row_idx


def clean_header(header_row):
    """清洗表头：去除 None，转为字符串并去除空白。"""
    result = []
    for v in header_row:
        if v is None:
            result.append('')
        else:
            result.append(str(v).strip())
    # 去除尾部空列
    while result and result[-1] == '':
        result.pop()
    return result


def clean_data_row(data_row, col_count):
    """清洗数据行，确保列数与表头一致。"""
    result = []
    for v in data_row[:col_count]:
        if v is None:
            result.append('')
        else:
            result.append(str(v).strip())
    # 补齐不足的列
    while len(result) < col_count:
        result.append('')
    return result


def main():
    src = os.path.join('data', 'basic_data.xlsx')
    out_dir = os.path.join('data', 'csv_output')
    os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.load_workbook(src)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sub_problems = parse_sub_problems(ws)

        if not sub_problems:
            print(f"[跳过] {sheet_name}: 未找到子问题")
            continue

        for name, header_row, data_rows in sub_problems:
            header = clean_header(header_row)
            if not header:
                print(f"[跳过] {name}: 表头为空")
                continue

            cleaned_data = [clean_data_row(r, len(header)) for r in data_rows]
            cleaned_data = [r for r in cleaned_data if any(c != '' for c in r)]

            out_path = os.path.join(out_dir, f'{name}.csv')
            with open(out_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(cleaned_data)

            print(f"[完成] {name}.csv — {len(cleaned_data)} 行数据")

    wb.close()
    print(f"\n所有 CSV 文件已输出到 {out_dir}")


if __name__ == '__main__':
    main()
